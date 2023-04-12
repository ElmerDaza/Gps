#documentacion oficial muy util: https://openpyxl.readthedocs.io/en/stable/tutorial.html
#from ast import Break
#from tkinter import font

import random
from openpyxl.styles import PatternFill
import openpyxl
import time
from sourse import funciones as fun, drive_operation as do
import os
from datetime import datetime as dt
import threading as hilo

#se obtiene los datos de fecha
day = dt.now()
dia = day.day
dia_str=dt.today().strftime('%A')
año=day.year
mes=day.month
mes_actual=mes


#variables para email
dic_DatosArchivoEmail={}


#variables
dic_dia_pago_user={}
dic_vehiculo_cliente={}
celdas_dia_anterior=None
dic_meses={'1':'enero',
            '2':'febrero','3':'marzo','4':'abril',
            '5':'mayo','6':'junio','7':'julio','8':'agosto',
            '9':'septiembre','10':'octubre','11':'noviembre',
            '12':'diciembre'
            }
# nombre de archivo
file_name=''
Nombre_Archivo=''
ID_file=''


#book=None
hoja= None
#abrir archivo
def open_file(direccion_archivo: str,nombre_hoja: str):
    #se recupera la llave
    directorio= os.path.split(direccion_archivo)[1]
    directorio= os.path.splitext(directorio)[0]
    with open('KEYS/'+directorio+'.key', 'rb') as filekey: 
        key = filekey.read()
    do.en.decryp(direccion_archivo,'',key)

    #acceder al libro
    book = openpyxl.load_workbook(direccion_archivo)
    #entrar a la hoja de interes
    hoja = book[nombre_hoja]
    
    #se encripta el archivo
    do.en.encryp(direccion_archivo)
    #se retorna la variable que contiene la informacion de la hoja
    return hoja, book


"""
#Obtener todas las hojas de trabajo
wb.sheetnames    #['sheet1']
 #Editar nombre de la hoja de trabajo
sheet.title="students"    
 #Obtener el nombre de la hoja de trabajo
sheet.title    #students"""

"""
optener una celda
sheet.cell(2,3)    #<Cell 'students'.C2>
sheet["C2"]    #<Cell 'students'.C2>
____________________
optener el valor
# método uno
sheet.cell(2,3).value    #60
 # Método dos
sheet["C2"].value    #60

cantidad de filas y columnas mas grandes
#Obtenga la fila más grande:
sheet.max_row    #8
 #Obtenga la columna más grande:
sheet.max_column    #4
___________________
mas instrucciones en: https://programmerclick.com/article/89381403512/
"""

def function_register():
    hoja,wb = open_file(file_name,'cobranza')
    max_row=hoja.max_row
    Cont=0
    info_completa=[]
    info_efimero=[0]
    for h in range(1, max_row+1):
        if hoja.cell(h,3)._value !=None and hoja.cell(h,3)._value !='None':
            Cont=Cont+1
        else:
            break
    max_row=Cont
    for i in (range(2,max_row+1)):
        
        info_efimero.append(hoja.cell(i,2)._value)
        info_efimero.append(hoja.cell(i,4)._value)
        if hoja.cell(i,5)._value == '' or hoja.cell(i,5)._value==None:
            info_efimero.append(str('ejemplo@elmerdaza.com'))
        else:
            info_efimero.append(hoja.cell(i,5)._value)
        
        if hoja.cell(i,6)._value == '' or hoja.cell(i,6)._value==None or len(hoja.cell(i,6)._value)<5:
            info_efimero.append(str(random.randrange(10000,19999,3)))
        else:
            info_efimero.append(hoja.cell(i,6)._value)
        info_efimero.append('')
        info_efimero.append(str(random.randrange(1010,9090,3)))
        info_efimero.append('')

        info_completa.append(info_efimero)
        info_efimero=[0]
    return info_completa
    



def procesar_datos():
    global dic_dia_pago_user
    global dic_vehiculo_cliente
    global dic_DatosArchivoEmail
    global celdas_dia_anterior
    #se abre el archivo y se optiene la hoja
    hoja, wb=open_file(file_name,"cobranza")
    #maximo de filas utilizadas

    max_row=hoja.max_row
    Cont=0
    for h in range(1, max_row+1):
        if hoja.cell(h,3)._value !=None and hoja.cell(h,3)._value !='None':
            Cont=Cont+1
        else:
            break
    max_row=Cont
    #maximo de columnas utilizadas
    max_colum = hoja.max_column
    #cantidad de celdas en uso
    cantidad_celdas = max_row*max_colum
    #variables contadoras
    contador_verde=0
    contador_rojo =0
    
    dic_telefono_cliente={}
    clientessinsumero=[]
    clients=[]
    months=[]

    
    
    
    print("filas: ",max_row,"columnas: ",max_colum,"total de celdas: ",cantidad_celdas)
    #time.sleep(2)
    #se recorre la cuadricula filas i columnas j
    for i in (range(2,max_row+1)):#filas
        #meses de deuda por cliente
        meses_deuda=0
        #llenar el diccionario
        #se optiene el nombre
        n=hoja.cell(i,2)._value
        #el nombre debe ser diferente de none
        if n!=None and n!="None":
            #se relaciona el cliente con su numero telefonico
            telefono=(str(hoja.cell(i,4)._value))
            dic_telefono_cliente.update({n:telefono})
            if telefono == "None" or telefono=="5252516":
                clientessinsumero.append(n)
            
        else:
            None

        #se recorren las columnas       
        for j in range(7,max_colum+1):#columnas
            #en la segunda columna esta los nombres de usuario
            if j == 7:#primera columna con informacion de pago
                nombre=hoja.cell(i,2)._value
                correo=hoja.cell(i,5)._value
                cedula=hoja.cell(i,6)._value
            #datos de la celda
            data = hoja.cell(i,j)
            #formato e informacion de la celda
            texto_fill = str(data.fill)
            #se divide el contenido en texto cada que aparezca ' 
            text_arr=texto_fill.split("'")
            #se recorre el arreglo con el texto y se compara la cadena
            #que mide exactamente 8 caracteres que es el codigo hexadecimal del color
            for element in text_arr:
                if 8 == len(element):
                    if element == "FF92D050":#color verde
                        #print("V se comparo: ", element)
                        contador_verde = contador_verde+1
                        break
                    elif element =='FFFF0000':#color rojo
                        #print("R se comparo: ", element)
                        contador_rojo =contador_rojo+1
                        meses_deuda=meses_deuda+1
                        break
        
        #dic_DatosArchivoEmail.update({nombre:[correo,cedula,telefono]})
        #cada vez que se completa una fila se agregan los datos
        clients.append(nombre)
        months.append(meses_deuda)




    
    #diccionario de cliente VS deuda
    #dic_usuarios_deuda= dict(zip(clients,months))

    #divicion de dias****************
    #variables
    dias=1
    contador_dias_filas={}
    
    #se recorre a partir de la segunda fila ya que
    #en la primera se encuentra los titulos
    for i in range(2,max_row+1):
        #informacion para cobro por gmail
        #________________________________
        name=hoja.cell(i,2)._value
        if dias==dia and name!=None and name!='None':
            #se agrega el nuevo nombre junto con la placa de esa fila
            name=str(name.strip())
            dic_vehiculo_cliente.update({name:[hoja.cell(i,3)._value]})
            next_=hoja.cell(i+1,2)._value
            #si es vacia quiere decir que pertenece al mismo usuario
            if next_==None:
                #se guarda el arreglo guardado
                
                arreglo=dic_vehiculo_cliente[name]
                #se agrega la nueva placa
                arreglo.append(hoja.cell(i+1,3)._value)
                for h in range(0,len(arreglo)):
                    if arreglo[h] == None or arreglo[h]=='None':
                        arreglo.pop(h)
                #y se actualiza e diccionario
                dic_vehiculo_cliente.update({name:arreglo})
                #a partir de aqui se recorren las siguientes celdas 
                #para saber si tambien pertenecen
                for num in range(i+2,max_row+1):
                    next_=hoja.cell(num,2)._value
                    if next_==None:
                        arreglo=dic_vehiculo_cliente[name]
                        arreglo.append(hoja.cell(num,3)._value)
                        for h in range(0,len(arreglo)):
                            if arreglo[h] == None or arreglo[h]=='None':
                                arreglo.pop(h)
                        dic_vehiculo_cliente.update({n:arreglo})
                    else:
                        break
            
            
            telefono=hoja.cell(i,4)._value
            correo=hoja.cell(i,5)._value
            cc=hoja.cell(i,6)._value
            if cc=='' or cc==None or cc=='5252516' or cc=='None':
                cc=str(random.randrange(1000000,1990099,3))
            fecha=f'{año}-{day.month}-{dias}'
            dic_DatosArchivoEmail.update({cc:[correo,name,telefono,fecha]})
        #__________________________________


        #se selecciona la celda de la columna 1
        celda = hoja.cell(i,1)._value
        
        #si es diferente de none se agrega un dato al contador
        if celda != None:
            dat=str(dias)
            efimero={dat:[i,0]}
            contador_dias_filas.update(efimero)
            #se agrega el dia 
            dic_dia_pago_user.update({str(dias):[clients[i-2]]})
        #al ser none se verifica la celda siguiente   
        else:
            # si es diferente de none 
            # quiere decir que hasta aqui contamos
            if hoja.cell(i+1,1)._value != None:
                dat=str(dias)
                d=contador_dias_filas.setdefault(dat)
                d[1]=i
                contador_dias_filas.update({str(dias):d})

                dic_dia=dic_dia_pago_user.setdefault(str(dias))
                dic_dia.append(clients[i-2])

                dias=dias+1
            else:#dia de pago por usuario
                dic_dia=dic_dia_pago_user.setdefault(str(dias))
                dic_dia.append(clients[i-2])
        #en el ultimo valor se agrega el dato faltante de la lista
        if i==max_row:
            dat=str(dias)
            d=contador_dias_filas.setdefault(dat)
            d[1]=max_row
            contador_dias_filas.update({str(dias):d})
    
    #si es lunes 1 se agrega en los usuarios a cobrar el ultimo dia del mes pasado
    if dia==1 and dia_str=='Monday':
        limites=contador_dias_filas[str(30)]
        #________________________________________________________
        for limite in range(limites[0],limites[1]+1):
            
            #_________________________________________________________________
            name=hoja.cell(limite,2)._value
            if name!=None and name!='None':
                
                #se agrega el nuevo nombre junto con la placa de esa fila
                name=str(name.strip())
                dic_vehiculo_cliente.update({name:[hoja.cell(limite,3)._value]})
                next_=hoja.cell(limite+1,2)._value
                #si es vacia quiere decir que pertenece al mismo usuario
                if next_==None:
                    #se guarda el arreglo guardado
                    
                    arreglo=dic_vehiculo_cliente[name]
                    #se agrega la nueva placa
                    arreglo.append(hoja.cell(limite+1,3)._value)
                    for h in range(0,len(arreglo)):
                        if arreglo[h] == None or arreglo[h]=='None':
                            arreglo.pop(h)
                    #y se actualiza e diccionario
                    dic_vehiculo_cliente.update({name:arreglo})
                    #a partir de aqui se recorren las siguientes celdas 
                    #para saber si tambien pertenecen
                    for num in range(limite+2,max_row+1):
                        next_=hoja.cell(num,2)._value
                        if next_==None:
                            arreglo=dic_vehiculo_cliente[name]
                            arreglo.append(hoja.cell(num,3)._value)
                            for h in range(0,len(arreglo)):
                                if arreglo[h] == None or arreglo[h]=='None':
                                    arreglo.pop(h)
                            dic_vehiculo_cliente.update({n:arreglo})
                        else:
                            break


                telefono=hoja.cell(limite,4)._value
                correo=hoja.cell(limite,5)._value
                cc=hoja.cell(limite,6)._value
                if cc=='' or cc==None or cc=='5252516' or cc=='None':
                    cc=str(random.randrange(1000000,1990099,3))
                
                if mes==1:
                    fecha=f'{año-1}-{12}-{30}'
                else:
                    fecha=f'{año}-{mes-1}-{30}'
                dic_DatosArchivoEmail.update({cc:[correo,name,telefono,fecha]})
    #si es lunes pero no 1 se agregan los usuarios del dia anterior
    elif dia_str=='Monday' and dia!=1:
        limites=contador_dias_filas[str(dia-1)]
        for limite in range(limites[0],limites[1]+1):
            

            #_______________________________________________________________
            name=hoja.cell(limite,2)._value
            if name!=None and name!='None':
                #se agrega el nuevo nombre junto con la placa de esa fila
                name=str(name.strip())
                dic_vehiculo_cliente.update({name:[hoja.cell(limite,3)._value]})
                next_=hoja.cell(limite+1,2)._value
                #si es vacia quiere decir que pertenece al mismo usuario
                if next_==None:
                    #se guarda el arreglo guardado
                    
                    arreglo=dic_vehiculo_cliente[name]
                    #se agrega la nueva placa
                    arreglo.append(hoja.cell(limite+1,3)._value)
                    for h in range(0,len(arreglo)):
                        if arreglo[h] == None or arreglo[h]=='None':
                            arreglo.pop(h)
                    #y se actualiza e diccionario
                    dic_vehiculo_cliente.update({name:arreglo})
                    #a partir de aqui se recorren las siguientes celdas 
                    #para saber si tambien pertenecen
                    for num in range(limite+2,max_row+1):
                        next_=hoja.cell(num,2)._value
                        if next_==None:
                            arreglo=dic_vehiculo_cliente[name]
                            arreglo.append(hoja.cell(num,3)._value)
                            for h in range(0,len(arreglo)):
                                if arreglo[h] == None or arreglo[h]=='None':
                                    arreglo.pop(h)
                            dic_vehiculo_cliente.update({n:arreglo})
                        else:
                            break
    
                telefono=hoja.cell(limite,4)._value
                correo=hoja.cell(limite,5)._value
                cc=hoja.cell(limite,6)._value
                if cc=='' or cc==None or cc=='5252516' or cc=='None':
                    cc=str(random.randrange(1000000,1990099,3))
                fecha=f'{año}-{mes}-{dia-1}'
                dic_DatosArchivoEmail.update({cc:[correo,name,telefono,fecha]})

    for d in dic_DatosArchivoEmail:
        arr=dic_DatosArchivoEmail[d]
        nam_=arr[1]
        arr.append(dic_vehiculo_cliente[nam_])
        dic_DatosArchivoEmail.update({d:arr})


    


    #MODIFICAR EL DOCUMENTO
    #numero de celdas del dia
    celdas_del_dia= contador_dias_filas[str(dia)]
    if dia!=1:
        celdas_dia_anterior=contador_dias_filas[str(dia-1)]
    else:
        celdas_dia_anterior=contador_dias_filas[str(30)]
    
    if dia_str=='Monday' and dia !=1:
    #se recorre las celdas correspondiente al dia
        for i in (range(celdas_dia_anterior[0],celdas_del_dia[1]+1)):#filas

            #variables
            #now=dt.now()
            num_celda_mes=0
            #mes=now.month
            mes_str=dic_meses[str(mes)]
            #veriable con los parametros de relleno en rojo
            red_fill = PatternFill(fill_type='solid', 
                                    start_color='00FF0000',
                                    end_color='00FF0000')
            
        
            #se selecciona la columna del mes
            for n in range(hoja.max_column+1):
                mes_hoja= hoja.cell(1,n+1)._value
                if mes_str==str(mes_hoja).lower():
                    num_celda_mes=n+1
                    break
            #se escoge la celda del respectivo usuario
            celda =hoja.cell(i,num_celda_mes)
            #los parametros de relleno se convierten a string
            fill_txt=str(celda.fill)
            #separacion por '
            fill_txt=fill_txt.split("'")

            
            for fill in fill_txt:
                if len(fill)==8:
                    print(fill)
                    if fill =='00000000':
                        celda.fill=red_fill
                        
                    break
    
    elif dia_str=='Monday' and dia ==1:
        for i in (range(celdas_dia_anterior[0],celdas_dia_anterior[1]+1)):#filas

            #variables
            #now=dt.now()
            num_celda_mes=0
            #mes=now.month
            if mes!=1:
                mes_str=dic_meses[str(mes-1)]
            else:
                mes_str=dic_meses[str(12)]
            #veriable con los parametros de relleno en rojo
            red_fill = PatternFill(fill_type='solid', 
                                    start_color='00FF0000',
                                    end_color='00FF0000')
            
        
            #se selecciona la columna del mes
            for n in range(hoja.max_column+1):
                mes_hoja= hoja.cell(1,n+1)._value
                if mes_str==str(mes_hoja).lower():
                    num_celda_mes=n+1
                    break
            #se escoge la celda del respectivo usuario
            celda =hoja.cell(i,num_celda_mes)
            #los parametros de relleno se convierten a string
            fill_txt=str(celda.fill)
            #separacion por '
            fill_txt=fill_txt.split("'")

            
            for fill in fill_txt:
                if len(fill)==8:
                    print(fill)
                    if fill =='00000000':
                        celda.fill=red_fill
                        
                    break
        
        for i in (range(celdas_del_dia[0],celdas_del_dia[1]+1)):#filas

            #variables
            #now=dt.now()
            num_celda_mes=0
            #mes=now.month
            mes_str=dic_meses[str(mes)]
            #veriable con los parametros de relleno en rojo
            red_fill = PatternFill(fill_type='solid', 
                                    start_color='00FF0000',
                                    end_color='00FF0000')
            
        
            #se selecciona la columna del mes
            for n in range(hoja.max_column+1):
                mes_hoja= hoja.cell(1,n+1)._value
                if mes_str==str(mes_hoja).lower():
                    num_celda_mes=n+1
                    break
            #se escoge la celda del respectivo usuario
            celda =hoja.cell(i,num_celda_mes)
            #los parametros de relleno se convierten a string
            fill_txt=str(celda.fill)
            #separacion por '
            fill_txt=fill_txt.split("'")

            
            for fill in fill_txt:
                if len(fill)==8:
                    print(fill)
                    if fill =='00000000':
                        celda.fill=red_fill
                        
                    break
    else:
        for i in (range(celdas_del_dia[0],celdas_del_dia[1]+1)):#filas
            #variables
            #now=dt.now()
            num_celda_mes=0
            #mes=now.month
            mes_str=dic_meses[str(mes)]
            #veriable con los parametros de relleno en rojo
            red_fill = PatternFill(fill_type='solid', 
                                    start_color='00FF0000',
                                    end_color='00FF0000')
            
            
            #se selecciona la columna del mes
            for n in range(hoja.max_column+1):
                mes_hoja= hoja.cell(1,n+1)._value
                if mes_str==str(mes_hoja).lower():
                    num_celda_mes=n+1
                    break
            #se escoge la celda del respectivo usuario
            celda =hoja.cell(i,num_celda_mes)
            #los parametros de relleno se convierten a string
            fill_txt=str(celda.fill)
            #separacion por '
            fill_txt=fill_txt.split("'")

            
            for fill in fill_txt:
                if len(fill)==8:
                    print(fill)
                    if fill =='00000000':
                        celda.fill=red_fill
                        
                    break    
    
    
    #se recupera la llave
    directorio= os.path.split(file_name)[1]
    directorio= os.path.splitext(directorio)[0]
    with open('KEYS/'+directorio+'.key', 'rb') as filekey: 
        key = filekey.read()
    #se desencripta el archivo
    do.en.decryp(file_name,"",key)
    #se guarda las modificaciones en el archivo descargado
    wb.save(file_name)
    #se actualiza el documento de drive
    guardar=hilo.Thread(target=do.actualizar_archivo,args=(Nombre_Archivo,file_name,ID_file))
    #do.actualizar_archivo(Nombre_Archivo,file_name,ID_file)
    guardar.start()
    




    #print(diccionario)
    #imprimir_diccionario(dic_vehiculo_cliente)
    #fun.imprimir_diccionario(dic_dia_pago_user)
    #print(dic_vehiculo_cliente['DELFIN PRIETO'])
    
    
    
    """
    ultimo_dia_por_mes={'1':31,
    '3':31,'4':30,'5':31,'6':30,'7':31,
    '8':31,'9':30,'10':31,'11':30,
    '12':31
    }
    
    if año%4 ==0:
        ultimo_dia_por_mes.update({'2':29})
    else:
        ultimo_dia_por_mes.update({'2':28})
    """
   


    #si es lunes se agregan los usuarios del dia anterior
    if dia_str == "Monday":
        #este metodo crea el archovo de cobro
        crear_txt_(
            dia,contador_dias_filas,clients,
            months,dic_telefono_cliente,
            "/sourse/texto/contactos_cobro.txt",contenido=[],lunes=True
            )
    else:#si no es lunes se cobra solo ese dia
        #este metodo crea el archovo de cobro
        crear_txt_(
            dia,contador_dias_filas,clients,
            months,dic_telefono_cliente,
            "/sourse/texto/contactos_cobro.txt",contenido=[]
            )
    #archivos de clientes sin numero
    crear_txt_(1,1,1,1,1,
        direccion_archivo="/sourse/texto/clientes_sin_numero.txt",
        contenido=clientessinsumero,
        Cobro=False
        )
    #fun.crear_archivo_texto(clientessinsumero,direccion)
    print('verdes: ',contador_verde)
    print('rojos: ',contador_rojo)
    

    #direccion del archivo
    address=os.getcwd()
    address+='/Archivos/'+Nombre_Archivo
    eliminar=hilo.Thread(target=eliminar_file(),args=(address))
    eliminar.start()
    
    
    

def eliminar_file(address:str):
    '''
    elimina el archivo de exel una vez que este confirmada la actualizacion
        address: ubicacion de archivo
    '''
    #verifica si existe el documento
    while os.path.exists(address):
        time.sleep(1)
        if fun.archivo_guardado:
            os.remove(file_name)

    
#funcion para crear archivo de texto
def crear_txt_(
    dia,dias_filas,clients,months,telefono_cliente,
    direccion_archivo,contenido,Cobro=True,lunes=False
    ):
    
    #si el archivo es para cobro entra aqui
    if Cobro:

        if lunes:#si es lunes entra aqui
            if dia!=1:
                #se agregan las celdas del dia anterior
                celdas_del_dia= dias_filas[str(dia)]
                completo=dias_filas[str(dia-1)]
                celdas_del_dia[0]=completo[0]
                inf_archivo=[]
                #impresion de datos
                print("HOY LES TOCA COBRO HA:",celdas_del_dia[0],celdas_del_dia[1])
                #confirmar si se cobra
                
                hoja, wb=open_file(file_name,"cobranza")
                #now=dt.now()
                #recorrer la filas y agregar los datos de cobro
                for numero in range(celdas_del_dia[0],celdas_del_dia[1]+1):
                    #variables
                    num_celda_mes=0
                    agregar_user=True
                    #mes=now.month
                    mes_str=dic_meses[str(mes)]
                    #se selecciona la columna del mes
                    for n in range(hoja.max_column+1):
                        mes_hoja= hoja.cell(1,n+1)._value
                        if mes_str==str(mes_hoja).lower():
                            num_celda_mes=n+1
                            break
                    #se escoge la celda del respectivo usuario
                    celda =hoja.cell(numero,num_celda_mes)
                    fill_txt=str(celda.fill)
                    fill_txt=fill_txt.split("'")
                    #se recorre el arreglo con el texto y se compara la cadena
                    #que mide exactamente 8 caracteres que es el codigo hexadecimal del color
                    for element in fill_txt:
                        if 8 == len(element):
                            if element == "FF92D050":#color verde
                                agregar_user=False
                                break
                            else:
                                break
                    
                    if agregar_user:
                        if clients[numero-2]!=None and clients[numero-2]!="None" and str(clients[numero-2]) !="5252516":
                            print(clients[numero-2],"-:-",numero)
                            dato=str(clients[numero-2])+": "+str(telefono_cliente[clients[numero-2]])
                            #deuda
                            for o in range(months[numero-2]):
                                dato+="-"
                            inf_archivo.append(dato)
                #crear archivo de texto
                
                direccion=os.getcwd()#direccion a la carpeta del proyecto
                direccion+="/sourse/texto/texto.txt"#se agrega la url adicional
                inf_=["Buen dia",
                "le escribo para informarle que su factura se generó, corespondiente al mes de {0}".format(mes_str),
                "mas una factura pendiente por pagar",
                "mas dos mensualidades vencidas, por lo tanto el servicio sera suspendido y si desea reactivarlo puede hacercarce a la oficina y cancelar el saldo pendiente.",
                ""
                ]
                fun.crear_archivo_texto(inf_,direccion)

                direccion=os.getcwd()#direccion a la carpeta del proyecto
                direccion+=direccion_archivo#se agrega la url adicional
                #en el scrip de funciones esta la funcion de archivo txt
                fun.crear_archivo_texto(inf_archivo,direccion)
            else:
                #se agregan las celdas del dia anterior
                celdas_del_dia= dias_filas[str(dia)]
                completo=dias_filas[str(dia)]
                celdas_del_dia[0]=completo[0]
                inf_archivo=[]
                
                #impresion de datos
                print("HOY LES TOCA COBRO HA:",celdas_del_dia[0],celdas_del_dia[1])
                
                #confirmar si se cobra
                hoja, wb=open_file(file_name,"cobranza")


                #recorrer la filas y agregar los datos de cobro
                for numero in range(celdas_del_dia[0],celdas_del_dia[1]+1):
                    #variables
                    num_celda_mes=0
                    agregar_user=True
                    #mes=now.month
                    mes_str=dic_meses[str(mes)]
                    #se selecciona la columna del mes
                    for n in range(hoja.max_column+1):
                        mes_hoja= hoja.cell(1,n+1)._value
                        if mes_str==str(mes_hoja).lower():
                            num_celda_mes=n+1
                            break
                    #se escoge la celda del respectivo usuario
                    celda =hoja.cell(numero,num_celda_mes)
                    fill_txt=str(celda.fill)
                    fill_txt=fill_txt.split("'")
                    #se recorre el arreglo con el texto y se compara la cadena
                    #que mide exactamente 8 caracteres que es el codigo hexadecimal del color
                    for element in fill_txt:
                        if 8 == len(element):
                            if element == "FF92D050":#color verde
                                agregar_user=False
                                break
                            else:
                                break
                    
                    if agregar_user:
                        if clients[numero-2]!=None and clients[numero-2]!="None" and str(clients[numero-2]) !="5252516":
                            print(clients[numero-2],"-:-",numero)
                            dato=str(clients[numero-2])+": "+str(telefono_cliente[clients[numero-2]])
                            #deuda
                            for o in range(months[numero-2]):
                                dato+="-"
                            inf_archivo.append(dato)
                #celdas_dia_anterior
                for numero in range(celdas_dia_anterior[0],celdas_dia_anterior[1]+1):
                    #variables
                    num_celda_mes=0
                    agregar_user=True
                    #mes=now.month
                    if mes!=1:
                        mes_str=dic_meses[str(mes-1)]
                    else:
                        mes_str=dic_meses[str(12)]
                    #se selecciona la columna del mes
                    for n in range(hoja.max_column+1):
                        mes_hoja= hoja.cell(1,n+1)._value
                        if mes_str==str(mes_hoja).lower():
                            num_celda_mes=n+1
                            break
                    #se escoge la celda del respectivo usuario
                    celda =hoja.cell(numero,num_celda_mes)
                    fill_txt=str(celda.fill)
                    fill_txt=fill_txt.split("'")
                    #se recorre el arreglo con el texto y se compara la cadena
                    #que mide exactamente 8 caracteres que es el codigo hexadecimal del color
                    for element in fill_txt:
                        if 8 == len(element):
                            if element == "FF92D050":#color verde
                                agregar_user=False
                                break
                            else:
                                break
                    
                    if agregar_user:
                        if clients[numero-2]!=None and clients[numero-2]!="None" and str(clients[numero-2]) !="5252516":
                            print(clients[numero-2],"-:-",numero)
                            dato=str(clients[numero-2])+": "+str(telefono_cliente[clients[numero-2]])
                            #deuda
                            for o in range(months[numero-2]):
                                dato+="-"
                            inf_archivo.append(dato)
                
                
                #crear archivo de texto
                

                direccion=os.getcwd()#direccion a la carpeta del proyecto
                direccion+="/sourse/texto/texto.txt"#se agrega la url adicional
                inf_=["Buen dia",
                "le escribo para informarle que su factura se generó, corespondiente al mes de {0}".format(mes_str),
                "mas una factura pendiente por pagar",
                "mas dos mensualidades vencidas, por lo tanto el servicio sera suspendido y si desea reactivarlo puede hacercarce a la oficina y cancelar el saldo pendiente.",
                ""
                ]
                fun.crear_archivo_texto(inf_,direccion)

                direccion=os.getcwd()#direccion a la carpeta del proyecto
                direccion+=direccion_archivo#se agrega la url adicional
                #en el scrip de funciones esta la funcion de archivo txt
                fun.crear_archivo_texto(inf_archivo,direccion)
        
        else:#si no es lunes entra aqui.
            #se escogen las filas
            celdas_del_dia= dias_filas[str(dia)]
            inf_archivo=[]
            #imprimir los datos
            print("HOY LES TOCA COBRO HA:",celdas_del_dia[0],celdas_del_dia[1])
            #recorrer las filas y almacenar los datos
            #confirmar si se cobra
            
            hoja, wb=open_file(file_name,"cobranza")
            #now=dt.now()
            #recorrer la filas y agregar los datos de cobro
            for numero in range(celdas_del_dia[0],celdas_del_dia[1]+1):
                #variables
                num_celda_mes=0
                agregar_user=True
                #mes=now.month
                mes_str=dic_meses[str(mes)]
                #se selecciona la columna del mes
                for n in range(hoja.max_column+1):
                    mes_hoja= hoja.cell(1,n+1)._value
                    if mes_str==str(mes_hoja).lower():
                        num_celda_mes=n+1
                        #print("pude ingresar-----------------------")
                        break
                #se escoge la celda del respectivo usuario
                celda =hoja.cell(numero,num_celda_mes)
                fill_txt=str(celda.fill)
                fill_txt=fill_txt.split("'")
                #se recorre el arreglo con el texto y se compara la cadena
                #que mide exactamente 8 caracteres que es el codigo hexadecimal del color
                for element in fill_txt:
                    if 8 == len(element):
                        if element == "FF92D050":#color verde
                            agregar_user=False
                            break
                        else:
                            break
                
                if agregar_user:
                    if clients[numero-2]!=None and clients[numero-2]!="None" and str(clients[numero-2]) !="5252516":
                        print(clients[numero-2],"-:-",numero)
                        dato=str(clients[numero-2])+": "+str(telefono_cliente[clients[numero-2]])
                        #deuda
                        for o in range(months[numero-2]):
                            dato+="-"
                        inf_archivo.append(dato)
            #crear archivo de texto
            #crear archivo de texto
            
            direccion=os.getcwd()#direccion a la carpeta del proyecto
            direccion+="/sourse/texto/texto.txt"#se agrega la url adicional
            inf_=["Buen dia",
            "le escribo para informarle que se genero su factura el dia {0} de {1}".format(dia,dic_meses[str(mes)]),
            "mas una factura pendiente por pagar",
            "mas dos mensualidades vencidas, por lo tanto el servicio sera suspendido y si desea reactivarlo puede hacercarce a la oficina y cancelar el saldo pendiente.",
            ""
            ]
            fun.crear_archivo_texto(inf_,direccion)



            direccion=os.getcwd()#direccion a la carpeta del proyecto
            direccion+=direccion_archivo#se agrega la url adicional
            #en el scrip de funciones esta la funcion de archivo txt
            fun.crear_archivo_texto(inf_archivo,direccion)
        
    else:
        #crear archivo de texto
        direccion=os.getcwd()
        direccion+=direccion_archivo
        fun.crear_archivo_texto(contenido,direccion)